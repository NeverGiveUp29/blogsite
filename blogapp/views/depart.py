from openpyxl import load_workbook

from django.shortcuts import render,redirect,HttpResponse

from blogapp import models

def depart_list_other(request):
    """部门列表"""
    return render(request, 'depart_list_other.html')

def depart_list(request):
    """部门列表"""
    # 去数据库获取所有的部门，返回对象是[对象，对象，对象]
    query_set = models.Department.objects.all()
    return render(request, 'depart_list.html', {'queryset':query_set})

def depart_add(request):
    """新建部门"""
    if request.method=="GET":
        # GET请求，直接返回到新增页面
        return render(request,'depart_add.html')
    elif request.method=="POST":
        # 获取部门名称
        title = request.POST.get("title")
        # 把页面表单获取的部门名称添加到数据库部门表
        models.Department.objects.create(title=title)
        return redirect('/depart/list/')

def depart_delete(request):
    """删除部门"""
    # 删除接口get请求：http://127.0.0.1:8000/depart/list/?nid=1
    # 获取id
    nid = request.GET.get("nid")
    # 数据库删除该部门
    # name = models.Department.objects.all()
    # print(name)
    # for i in name:
    #     print(i.id,i.title)
    models.Department.objects.filter(id=nid).delete()
    # 删除完返回到列表页面
    # return HttpResponse("删除成功！")
    return redirect('/depart/list/')

def depart_edit(request,nid):
    """编辑部门"""
    if request.method=="GET":
        # 通过界面传回的url获取nid，然后从表中查到该id对应的title,最后回填到编辑框里
        depart_obj = models.Department.objects.filter(id=nid).first()
        return render(request,"depart_edit.html",{"depart_obj":depart_obj})
    # 提交表单时进行以下操作
    # 获取提交的部门名称
    title = request.POST.get("title")
    # 数据库更新部门名称
    models.Department.objects.filter(id=nid).update(title=title)
    # 重定向回列表页面
    return redirect('/depart/list/')

def depart_upload(request):
    """批量新建部门"""
    if request.method=='get':
        return render(request,'depart_edit.html')
    file_object = request.FILES.get('uploadFile')
    # print(file_object)# -->批量导入部门.xlsx
    if file_object:
        wb = load_workbook(file_object)
        sheet = wb.worksheets[0]
        for cell in sheet.iter_rows(min_row=2):
            # print(cell[0].value)# -->策划部
            exists = models.Department.objects.filter(title=cell[0].value).exists()
            if exists:
                pass
            models.Department.objects.create(title=cell[0].value)
        return redirect('/depart/list/')
    return HttpResponse('上传文件不能为空！')